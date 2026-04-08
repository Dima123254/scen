# -*- coding: utf-8 -*-
"""Генерация отчётов валидации и таблиц приложений ФСТЭК."""

from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from core.models import CompanyData
from reporting import fstec_appendices
from validation.filter_engine import FilterEngineReport

logger = logging.getLogger(__name__)


class ReportGenerator:
    """TXT/JSON/XLSX отчёты и приложения 6–11."""

    def generate_validation_report(
        self,
        engine_report: FilterEngineReport,
        out_path: Path,
        formats: list[str] | None = None,
    ) -> list[Path]:
        """Отчёт о валидации со статистикой и причинами отклонения.

        Args:
            engine_report: Результат ``FilterEngine.run``.
            out_path: Базовый путь без расширения или каталог.
            formats: Подмножество ``json``, ``txt``, ``xlsx``.

        Returns:
            Список созданных файлов.
        """
        formats = formats or ["json", "txt"]
        out_path = Path(out_path)
        if out_path.suffix:
            base = out_path.with_suffix("")
        else:
            base = out_path / "validation_report"
            base.parent.mkdir(parents=True, exist_ok=True)

        written: list[Path] = []
        payload: dict[str, Any] = {
            "summary_by_filter": engine_report.summary_by_filter,
            "scenarios": engine_report.per_scenario,
        }
        if "json" in formats:
            p = Path(str(base) + ".json")
            p.parent.mkdir(parents=True, exist_ok=True)
            p.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
            written.append(p)
        if "txt" in formats:
            p = Path(str(base) + ".txt")
            p.parent.mkdir(parents=True, exist_ok=True)
            lines = ["=== Отчёт валидации сценариев ===", ""]
            lines.append("Сводка по фильтрам:")
            for fname, stat in engine_report.summary_by_filter.items():
                lines.append(
                    f"  {fname}: пройдено {stat.get('passed', 0)}, "
                    f"отклонено {stat.get('failed', 0)}"
                )
            lines.append("")
            for row in engine_report.per_scenario:
                lines.append(f"Сценарий: {row['scenario_id']}")
                lines.append(f"  Все фильтры: {'ДА' if row['all_passed'] else 'НЕТ'}")
                for f in row["filters"]:
                    st = "OK" if f["passed"] else "FAIL"
                    lines.append(f"    [{st}] {f['name']}: {f['reason']}")
                lines.append("")
            p.write_text("\n".join(lines), encoding="utf-8")
            written.append(p)
        if "xlsx" in formats:
            p = Path(str(base) + ".xlsx")
            p.parent.mkdir(parents=True, exist_ok=True)
            wb = Workbook()
            ws = wb.active
            ws.title = "summary"
            ws.append(["filter", "passed", "failed"])
            for k, v in engine_report.summary_by_filter.items():
                ws.append([k, v.get("passed", 0), v.get("failed", 0)])
            ws2 = wb.create_sheet("details")
            ws2.append(["scenario_id", "filter", "passed", "reason"])
            for row in engine_report.per_scenario:
                for f in row["filters"]:
                    ws2.append([row["scenario_id"], f["name"], f["passed"], f["reason"]])
            wb.save(p)
            written.append(p)
        return written

    def generate_fstec_appendix(
        self,
        appendix_number: int,
        company: CompanyData,
        out_path: Path,
        fmt: str = "json",
    ) -> Path:
        """Экспорт таблицы приложения 6–11.

        Args:
            appendix_number: 6–11.
            company: Данные организации.
            out_path: Файл или каталог.
            fmt: ``json``, ``txt``, ``xlsx``.

        Returns:
            Путь к файлу.
        """
        title, rows = fstec_appendices.build_appendix(appendix_number, company)
        out_path = Path(out_path)
        if out_path.is_dir() or str(out_path).endswith("\\") or str(out_path).endswith("/"):
            out_path = out_path / f"appendix_{appendix_number}.{fmt}"
        if fmt == "json":
            out_path.parent.mkdir(parents=True, exist_ok=True)
            out_path.write_text(
                json.dumps({"title": title, "rows": rows}, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            return out_path
        if fmt == "txt":
            out_path.parent.mkdir(parents=True, exist_ok=True)
            if not rows:
                text = title + "\n(нет строк)\n"
            else:
                keys = list(rows[0].keys())
                text = title + "\n" + "; ".join(keys) + "\n"
                for r in rows:
                    text += "; ".join(str(r[k]) for k in keys) + "\n"
            out_path.write_text(text, encoding="utf-8")
            return out_path
        if fmt == "xlsx":
            return self.export_to_xlsx(rows, out_path, sheet_name=f"A{appendix_number}")
        raise ValueError(f"Unsupported format: {fmt}")

    def export_to_xlsx(
        self,
        rows: list[dict[str, Any]],
        out_path: Path,
        sheet_name: str = "Sheet1",
    ) -> Path:
        """Экспорт произвольных строк в Excel.

        Args:
            rows: Список однотипных словарей.
            out_path: Путь к .xlsx.
            sheet_name: Имя листа.

        Returns:
            Путь к файлу.
        """
        out_path = Path(out_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]
        if not rows:
            wb.save(out_path)
            return out_path
        keys = list(rows[0].keys())
        ws.append(keys)
        for r in rows:
            ws.append([r.get(k) for k in keys])
        wb.save(out_path)
        return out_path
